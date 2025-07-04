import pandas as pd
import os
from datetime import datetime
# تم حذف جميع دوال وتقنيات تصدير PDF نهائياً بناءً على طلب المستخدم.

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def export_cases_to_excel(cases, filename, title="تقرير الحالات", custom_columns=None):
    """
    تصدير الحالات إلى ملف Excel بتنسيق احترافي.
    cases: قائمة dicts
    filename: اسم الملف
    title: عنوان التقرير (اختياري)
    custom_columns: قائمة أعمدة مخصصة [("اسم العرض", "مفتاح الدكت")]
    """
    try:
        if not cases or not isinstance(cases, list) or not cases or not isinstance(cases[0], dict):
            # بيانات فارغة أو غير متوافقة
            import pandas as pd
            df = pd.DataFrame()
            df.to_excel(filename, index=False)
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "الحالات"
        # إعداد أنماط التنسيق
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        title_font = Font(bold=True, size=14)
        align_center = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        # إعداد الأعمدة
        if custom_columns:
            columns = custom_columns
        else:
            columns = [(k, k) for k in cases[0].keys()]
        # إضافة العنوان
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=1).alignment = align_center
        # رؤوس الأعمدة
        for col_idx, (col_name, col_key) in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border
        # البيانات
        for row_idx, row in enumerate(cases, 3):
            for col_idx, (col_name, col_key) in enumerate(columns, 1):
                value = row.get(col_key, "") if isinstance(row, dict) else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = align_center
                cell.border = border
        # ضبط عرض الأعمدة تلقائياً
        for col_idx, (col_name, col_key) in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row in cases:
                val = row.get(col_key, "") if isinstance(row, dict) else ""
                max_length = max(max_length, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_length + 2, 40))
        wb.save(filename)
    except Exception as e:
        raise Exception(f"فشل في حفظ ملف Excel: {e}")
